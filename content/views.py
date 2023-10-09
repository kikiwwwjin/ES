from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import Model_result
import simplejson as json
import os
import openpyxl
import pandas as pd
import numpy as np
# Create your views here.

class Sk_magic(APIView):
    def get(self, request):
        print('sk magic 모델 get으로 호출')
        result_dict = {
            'site' : 'sk_magic'
        }
        return render(request, 'models/sk_magic.html', result_dict)

    def post(self, request):
        print('sk magic 모델 post로 호출')
        return render(request, 'models/sk_magic.html')

class UploadData(APIView):
    def get(self, request):

        # get_data = request.data.get('ti_op_submit')
        # print(f'받은 데이터 : {get_data}')

        # 보내기 view 단으로 Rendering
        return Response(status=200) # context 딕셔너리 형태

    def post(self, request):
        title = request.POST.get('title')
        option_name = request.POST.get('option_name')
        ti_op_dict = {'title': title, 'option_name': option_name}
        print('view단에 왔다!!', ti_op_dict)
        # get_data = request.POST.get('ti_op_submit')
        # print(f'받은 데이터 : {get_data}')
        '''
        AJAX 통신을 통해 가져온 타이틀/옵션명을 정제(모델 결과)
        '''
        # 분석 프로젝트 경로 세팅
        back_pjt_path = 'C:/Users/kjh/PycharmProjects/pythonProject2/'
        os.system('conda info --envs') # 가상환경 확인
        os.system(f'python {back_pjt_path}main_v2.py --code=1 --site=sk_magic --title={title} --option_name={option_name}')
        print('모델결과 DB에 적재 완료!!')

        # 모델 결과값 DB에서 불러오기(가장 최근에 INSERT된 값, 필드명 : id)
        model_result_data = Model_result.objects.last()
        print(model_result_data.manufacturer)

        model_result_data = {
            'manufacturer': model_result_data.manufacturer,
            'brand': model_result_data.brand,
            'model_name': model_result_data.model_name,
        }
        print(model_result_data)

        # data_dict = [{'manu' : data.manufacturer} for data in model_result_data]
        # print(data_dict)
        # 보내기 view 단으로 Rendering
        # return Response(status=200) # context 딕셔너리 형태
        # Return 값 넘겨주기
        return JsonResponse(model_result_data)
        # return Response(status=200)
        # return HttpResponse(json.dumps(dict(model_result_data=model_result_data)), content_type='application/json')

class Upload_item_master(APIView):
    def get(self, request):
        print('Item_master_upload get으로 호출')
        return render(request, 'models/sk_magic.html')

    def post(self, request):
        print('Item_master_upload post로 호출')
        excel_file_object = request.FILES['files']
        site = request.POST.get['site']
        site_2 = request.FILES['site']
        # excel_file_object = request.POST.get['item_master']
        print(f'엑셀 파일 객체 : {excel_file_object}')
        print(f'사이트(프로젝트) 객체 : {site}')
        print(f'site_2 객체 : {site_2}')

        # AJAX로 가져온 엑셀 파일 객체 python에서 읽어 들이기
        excel = openpyxl.load_workbook(excel_file_object, data_only=True)
        # excel =openpyxl.load_workbook(excel_file_object)
        work_sheet = excel.worksheets[0] # 첫번째 시트 가져오기
        print(f'엑셀 파일 첫번째 시트 : {work_sheet}')

        # 엑셀 시트의 데이터 가져오기
        all_values = list()
        columns_list = list()
        for i, row in enumerate(work_sheet.rows):
            row_value = []
            for cell in row:
                if i == 0:
                    columns_list.append(cell.value)
                else:
                    # print(f'row 정보 : {row} || cel 정보 : {cell.value}')
                    row_value.append(cell.value)
            # 행 기준 value 값 APPEND
            if row_value != []:
                all_values.append(row_value)

        # DataFrame으로 변환
        df = pd.DataFrame(data=all_values, columns=columns_list)
        print(df)




        # DB 정보 가져오기
        # print(f'업로드 할 DB 테이블 정보 : {}')


        # 해당 테이블로 업로드



        # print(f'{} 아이템 마스터 업로드가 완료되었습니다.')
        # 아이템 마스터 객체 처리


        # 결과 변수 리턴
        # result = '아이템 마스터 객체'
        response = {'status':1, 'message':'정상적으로 작동'}
        # return HttpResponse(json.dumps(response), content_type='application/json')
        return JsonResponse(response)

class Dupcheck_item_master(APIView):
    def get(self, request):
        print('Item_master_upload get으로 호출')
        return render(request, 'models/sk_magic.html')

    def post(self, request):
        print('Dupcheck_item_master post로 호출')
        item_master_obj = request.FILES['item_master']  # 아이템 마스터
        selected_columns_obj = request.POST.get('selected_columns')  # 선택된 중복제거 컬럼 리스트
        print(f'AJAX로 가지고 온 아이템 마스터 객체 : {item_master_obj}')
        print(f'AJAX로 가지고 온 중복제거 컬럼 객체 : {selected_columns_obj}')

        # 1. 아이템 마스터
        excel = openpyxl.load_workbook(item_master_obj)
        print(f'해당 엑셀 시트 확인 : {excel.worksheets}')
        work_sheet = excel.worksheets[0]  # 첫번째 시트 가져오기
        print(f'엑셀 파일 첫번째 시트 : {work_sheet}')

        # 2. 컬럼 리스트화
        selected_columns_list = selected_columns_obj.split(',')

        # 엑셀 시트의 데이터 가져오기
        all_values = list()
        columns_list = list()
        for i, row in enumerate(work_sheet.rows):
            row_value = []
            for cell in row:
                if i == 0:
                    columns_list.append(cell.value)
                else:
                    # print(f'row 정보 : {row} || cel 정보 : {cell.value}')
                    row_value.append(cell.value)
            # 행 기준 value 값 APPEND
            if row_value != []:
                all_values.append(row_value)

        # DataFrame으로 변환
        df = pd.DataFrame(data=all_values, columns=columns_list)
        df = df.fillna('')
        print(f'해당 데이터 \n {df}')

        # 3. 중복 확인
        dup_df = df[df.duplicated(selected_columns_list, keep=False)]
        # dup_df = dup_df[dup_df.duplicated(selected_columns_list, keep=False)]
        print(f'확인된 중복 건수 : {dup_df}')

        # dup_df = pd.DataFrame({
        #     '이름':['김진형','김진형','오하람'],
        #     '나이':[31,31,33],
        # })

        # 4. 결과 JSON 생성
        response = dict()
        if (len(dup_df) > 0):
            print('중복 존재')
            dup_df = dup_df.reset_index(drop=True)
            dup_dict = dict()
            for col in selected_columns_list:
                dup_dict[col] = dup_df[col].tolist()

            # response 딕셔너리에 결과 적재
            response['dup_check'] = '중복 존재'
            response['dup_data'] = dup_dict
        else:
            response['dup_check'] = '중복 없음'
            response['dup_data'] = {}

        # fs = FileSystemStorage()
        # print('엑셀 파일 객체 :', excel_file)
        #
        #
        # # filename = fs.save('/data/file_upload/test.xlsx', file)
        # filename = fs.save('C:/Users/Admin/Documents/GitHub/ES/test.xlsx', file)
        # print(f'filename : {filename}')
        # print(f'{file} 적재가 완료되었습니다!!')

        # 아이템 마스터 객체 처리

        # 결과 변수 리턴
        # result = '아이템 마스터 객체'
        # response = {'status':1, 'message':'정상적으로 작동'}
        # return HttpResponse(json.dumps(response), content_type='application/json')
        return JsonResponse(response)

class Filecheck_item_master(APIView):
    def get(self, request):
        print('Item_master_upload get으로 호출')
        return render(request, 'models/sk_magic.html')

    def post(self, request):
        print('Item_master_upload post로 호출')
        excel_file_object = request.FILES['files']
        # excel_file_object = request.POST.get['item_master']
        print(f'엑셀 파일 객체 : {excel_file_object}')

        # AJAX로 가져온 엑셀 파일 객체 python에서 읽어 들이기
        excel = openpyxl.load_workbook(excel_file_object, data_only=True)
        # excel =openpyxl.load_workbook(excel_file_object)
        work_sheet = excel.worksheets[0] # 첫번째 시트 가져오기
        print(f'엑셀 파일 첫번째 시트 : {work_sheet}')

        # 엑셀 시트의 데이터 가져오기
        all_values = list()
        columns_list = list()
        for i, row in enumerate(work_sheet.rows):
            row_value = []
            for cell in row:
                if i == 0:
                    columns_list.append(cell.value)
                else:
                    # print(f'row 정보 : {row} || cel 정보 : {cell.value}')
                    row_value.append(cell.value)
            # 행 기준 value 값 APPEND
            if row_value != []:
                all_values.append(row_value)

        # DataFrame으로 변환
        df = pd.DataFrame(data=all_values, columns=columns_list)
        print('데이터 :', df)




        # 컬럼 정보 Dictionary에 담기
        data_info_dict = {'column_list' : columns_list}


        return JsonResponse(data_info_dict)